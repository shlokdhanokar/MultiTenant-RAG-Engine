"""
XLSX layout parser.

Spreadsheets have no heading hierarchy, so the heading-based strategy used for
PDF/DOCX/PPTX doesn't apply. Instead each sheet becomes a Topic and its rows
are serialized as "Header: value" pairs — that phrasing keeps the column name
attached to the value inside the chunk text, which is what makes a row
retrievable by a natural-language question about one of its fields.
"""
import logging

logger = logging.getLogger(__name__)

# Guardrail against pathological spreadsheets producing tens of thousands of chunks.
MAX_ROWS_PER_SHEET = 5000


def _looks_like_header(row):
    """A header row is all-text and fully populated."""
    values = [c for c in row if c is not None and str(c).strip()]
    if not values or len(values) < len([c for c in row if c is not None]):
        return False
    return all(isinstance(c, str) for c in values)


def analyze_xlsx_layout(xlsx_bytes: bytes):
    """
    Extracts sheet-scoped blocks from an .xlsx file.
    Embedded images are not extracted — spreadsheet charts/images rarely carry
    standalone meaning without their surrounding cells.
    """
    import io
    from openpyxl import load_workbook

    try:
        # read_only keeps memory flat on large files; data_only returns computed
        # values instead of formula strings, which is what we want to index.
        workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"failed to load xlsx from memory: {e}")
        return None

    semantic_blocks = []

    for sheet_index, sheet in enumerate(workbook.worksheets):
        semantic_blocks.append({
            "page_number": sheet_index,
            "text": sheet.title,
            "font_size": 0,
            "bbox": (0, 0, 0, 0),
            "type": "Topic",
        })

        headers = None
        row_position = 1
        truncated = False

        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_index >= MAX_ROWS_PER_SHEET:
                truncated = True
                break

            if row is None or all(c is None or not str(c).strip() for c in row):
                continue

            if headers is None and _looks_like_header(row):
                headers = [str(c).strip() if c is not None else "" for c in row]
                continue

            if headers:
                pairs = [
                    f"{headers[i]}: {value}"
                    for i, value in enumerate(row)
                    if i < len(headers) and value is not None and str(value).strip() and headers[i]
                ]
                text = " | ".join(pairs)
            else:
                text = " | ".join(str(c).strip() for c in row if c is not None and str(c).strip())

            if not text:
                continue

            semantic_blocks.append({
                "page_number": sheet_index,
                "text": text,
                "font_size": 0,
                "bbox": (0, row_position, 0, row_position),
                "type": "Paragraph",
            })
            row_position += 1

        if truncated:
            logger.warning(
                f"sheet '{sheet.title}' exceeded {MAX_ROWS_PER_SHEET} rows; remaining rows were skipped"
            )

    workbook.close()

    return {
        "total_pages": len(workbook.worksheets),
        "semantic_blocks": semantic_blocks,
        "images": [],
    }
